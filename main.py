from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates
from datetime import datetime
from typing import Optional, List
import uvicorn
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

app = FastAPI(title="OYE Hotels API", version="1.0.0")

# Setup templates
templates = Jinja2Templates(directory="templates")

# Comprehensive hotel data with real Unsplash images
hotels = [
    # Mumbai Hotels
    {
        "id": 1,
        "name": "OYE Grand Palace",
        "location": "Mumbai",
        "city": "Mumbai",
        "price": 3500,
        "rating": 4.5,
        "reviews": 256,
        "image": "https://images.unsplash.com/photo-1566073771259-6a8506099945?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Pool", "Restaurant", "Gym", "Room Service"],
        "description": "Luxury hotel in the heart of Mumbai with stunning city views",
        "type": "Luxury"
    },
    {
        "id": 2,
        "name": "OYE Marine Drive Suites",
        "location": "Mumbai",
        "city": "Mumbai",
        "price": 4500,
        "rating": 4.8,
        "reviews": 189,
        "image": "https://images.unsplash.com/photo-1582719508461-905c673771fd?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Sea View", "Restaurant", "Bar", "Spa"],
        "description": "Premium suites overlooking Marine Drive",
        "type": "Luxury"
    },

    # Bengaluru Hotels
    {
        "id": 3,
        "name": "OYE Tech Park Residency",
        "location": "Bengaluru",
        "city": "Bengaluru",
        "price": 2800,
        "rating": 4.6,
        "reviews": 324,
        "image": "https://images.unsplash.com/photo-1542314831-068cd1dbfeeb?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Business Center", "Gym", "Parking", "Restaurant"],
        "description": "Modern hotel near IT parks, perfect for business travelers",
        "type": "Business"
    },
    {
        "id": 4,
        "name": "OYE Garden Retreat",
        "location": "Bengaluru",
        "city": "Bengaluru",
        "price": 3200,
        "rating": 4.7,
        "reviews": 198,
        "image": "https://images.unsplash.com/photo-1520250497591-112f2f40a3f4?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Garden View", "Pool", "Spa", "Restaurant"],
        "description": "Peaceful garden hotel in Whitefield area",
        "type": "Resort"
    },

    # Hyderabad Hotels
    {
        "id": 5,
        "name": "OYE Charminar Heritage",
        "location": "Hyderabad",
        "city": "Hyderabad",
        "price": 2500,
        "rating": 4.4,
        "reviews": 167,
        "image": "https://images.unsplash.com/photo-1571896349842-33c89424de2d?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Heritage Property", "Restaurant", "Parking", "Room Service"],
        "description": "Heritage hotel near historic Charminar",
        "type": "Heritage"
    },
    {
        "id": 6,
        "name": "OYE Hi-Tech City Plaza",
        "location": "Hyderabad",
        "city": "Hyderabad",
        "price": 3000,
        "rating": 4.6,
        "reviews": 234,
        "image": "https://images.unsplash.com/photo-1564501049412-61c2a3083791?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Business Center", "Gym", "Pool", "Restaurant"],
        "description": "Contemporary hotel in Hi-Tech City business district",
        "type": "Business"
    },

    # Chennai Hotels
    {
        "id": 7,
        "name": "OYE Marina Beach Resort",
        "location": "Chennai",
        "city": "Chennai",
        "price": 3300,
        "rating": 4.7,
        "reviews": 289,
        "image": "https://images.unsplash.com/photo-1551882547-ff40c63fe5fa?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Beach Access", "Pool", "Restaurant", "Spa"],
        "description": "Beachfront resort with stunning Marina views",
        "type": "Resort"
    },
    {
        "id": 8,
        "name": "OYE T Nagar Grand",
        "location": "Chennai",
        "city": "Chennai",
        "price": 2600,
        "rating": 4.5,
        "reviews": 156,
        "image": "https://images.unsplash.com/photo-1578683010236-d716f9a3f461?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Restaurant", "Gym", "Parking", "Room Service"],
        "description": "Prime location in T Nagar shopping district",
        "type": "Standard"
    },

    # Delhi Hotels
    {
        "id": 9,
        "name": "OYE Connaught Plaza",
        "location": "Delhi",
        "city": "Delhi",
        "price": 3500,
        "rating": 4.6,
        "reviews": 312,
        "image": "https://images.unsplash.com/photo-1596436889106-be35e843f974?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Business Center", "Restaurant", "Bar", "Gym"],
        "description": "Central location near Connaught Place",
        "type": "Business"
    },
    {
        "id": 10,
        "name": "OYE Airport Suites",
        "location": "Delhi",
        "city": "Delhi",
        "price": 2900,
        "rating": 4.4,
        "reviews": 178,
        "image": "https://images.unsplash.com/photo-1611892440504-42a792e24d32?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Airport Shuttle", "Restaurant", "Gym", "Parking"],
        "description": "Convenient hotel near IGI Airport",
        "type": "Business"
    },

    # Goa Hotels
    {
        "id": 11,
        "name": "OYE Beach Paradise",
        "location": "Goa",
        "city": "Goa",
        "price": 4200,
        "rating": 4.9,
        "reviews": 445,
        "image": "https://images.unsplash.com/photo-1568084680786-a84f91d1153c?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Beach Access", "Pool", "Bar", "Spa", "Water Sports"],
        "description": "Luxury beach resort in North Goa",
        "type": "Resort"
    },
    {
        "id": 12,
        "name": "OYE Coastal Retreat",
        "location": "Goa",
        "city": "Goa",
        "price": 3800,
        "rating": 4.7,
        "reviews": 267,
        "image": "https://images.unsplash.com/photo-1571003123894-1f0594d2b5d9?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Beach View", "Pool", "Restaurant", "Spa"],
        "description": "Serene beachside property in South Goa",
        "type": "Resort"
    },

    # Jaipur Hotels
    {
        "id": 13,
        "name": "OYE Royal Heritage",
        "location": "Jaipur",
        "city": "Jaipur",
        "price": 3100,
        "rating": 4.8,
        "reviews": 223,
        "image": "https://images.unsplash.com/photo-1455587734955-081b22074882?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Heritage Property", "Pool", "Restaurant", "Cultural Programs"],
        "description": "Restored palace hotel near City Palace",
        "type": "Heritage"
    },

    # Pune Hotels
    {
        "id": 14,
        "name": "OYE Koregaon Park",
        "location": "Pune",
        "city": "Pune",
        "price": 2700,
        "rating": 4.5,
        "reviews": 189,
        "image": "https://images.unsplash.com/photo-1542314831-068cd1dbfeeb?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Restaurant", "Gym", "Parking", "Pool"],
        "description": "Trendy hotel in Koregaon Park area",
        "type": "Standard"
    },

    # Kolkata Hotels
    {
        "id": 15,
        "name": "OYE Park Street Grand",
        "location": "Kolkata",
        "city": "Kolkata",
        "price": 2800,
        "rating": 4.6,
        "reviews": 198,
        "image": "https://images.unsplash.com/photo-1566073771259-6a8506099945?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Restaurant", "Bar", "Gym", "Room Service"],
        "description": "Classic hotel on iconic Park Street",
        "type": "Standard"
    },

    # Udaipur Hotels
    {
        "id": 16,
        "name": "OYE Lake Palace View",
        "location": "Udaipur",
        "city": "Udaipur",
        "price": 4500,
        "rating": 4.9,
        "reviews": 356,
        "image": "https://images.unsplash.com/photo-1582719508461-905c673771fd?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "Lake View", "Pool", "Restaurant", "Spa", "Boat Rides"],
        "description": "Luxury hotel overlooking Lake Pichola",
        "type": "Luxury"
    },

    # Budget Options
    {
        "id": 17,
        "name": "OYE Budget Stay Mumbai",
        "location": "Mumbai",
        "city": "Mumbai",
        "price": 1500,
        "rating": 4.2,
        "reviews": 423,
        "image": "https://images.unsplash.com/photo-1520250497591-112f2f40a3f4?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "AC", "Parking"],
        "description": "Affordable and clean budget hotel",
        "type": "Budget"
    },
    {
        "id": 18,
        "name": "OYE Budget Stay Bengaluru",
        "location": "Bengaluru",
        "city": "Bengaluru",
        "price": 1800,
        "rating": 4.3,
        "reviews": 367,
        "image": "https://images.unsplash.com/photo-1551882547-ff40c63fe5fa?w=800&h=600&fit=crop",
        "amenities": ["WiFi", "AC", "Parking", "Breakfast"],
        "description": "Value for money accommodation",
        "type": "Budget"
    },
]

# Store bookings (in-memory for demo)
bookings = []

# Excel file path
EXCEL_FILE = "bookings.xlsx"

# Initialize Excel file
def init_excel_file():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hotel Bookings"

        # Headers
        headers = ["Booking ID", "Hotel Name", "Hotel Location", "Hotel Type", "Guest Name",
                   "Email", "Phone", "Check-in Date", "Check-out Date", "Number of Guests",
                   "Price per Night", "Booking Date & Time"]

        # Style headers
        header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 20

        wb.save(EXCEL_FILE)
        print(f"✅ Excel file created: {EXCEL_FILE}")

def save_booking_to_excel(booking):
    """Save a single booking to Excel file"""
    try:
        # Load existing workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        # Add booking data
        hotel = booking["hotel"]
        ws.cell(row=next_row, column=1).value = booking["id"]
        ws.cell(row=next_row, column=2).value = hotel["name"]
        ws.cell(row=next_row, column=3).value = hotel["location"]
        ws.cell(row=next_row, column=4).value = hotel.get("type", "N/A")
        ws.cell(row=next_row, column=5).value = booking["guest_name"]
        ws.cell(row=next_row, column=6).value = booking["email"]
        ws.cell(row=next_row, column=7).value = booking["phone"]
        ws.cell(row=next_row, column=8).value = booking["check_in"]
        ws.cell(row=next_row, column=9).value = booking["check_out"]
        ws.cell(row=next_row, column=10).value = booking["guests"]
        ws.cell(row=next_row, column=11).value = f"₹{hotel['price']}"
        ws.cell(row=next_row, column=12).value = booking["booking_date"]

        # Center align all cells in the new row
        for col in range(1, 13):
            ws.cell(row=next_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Alternate row colors
        if next_row % 2 == 0:
            fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            for col in range(1, 13):
                ws.cell(row=next_row, column=col).fill = fill

        wb.save(EXCEL_FILE)
        print(f"✅ Booking #{booking['id']} saved to Excel")
        return True
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")
        return False

# Initialize Excel file on startup
init_excel_file()

# Get unique cities for autocomplete
def get_cities():
    return sorted(list(set(hotel["city"] for hotel in hotels)))

# Get unique amenities
def get_all_amenities():
    amenities_set = set()
    for hotel in hotels:
        hotel_amenities = hotel.get("amenities", [])
        if isinstance(hotel_amenities, list):
            amenities_set.update(hotel_amenities)
    return sorted(list(amenities_set))

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    cities = get_cities()
    amenities = get_all_amenities()
    return templates.TemplateResponse("index.html", {
        "request": request,
        "hotels": hotels,
        "cities": cities,
        "amenities": amenities
    })

@app.post("/api/book")
async def book_hotel(
    hotel_id: int = Form(...),
    guest_name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    check_in: str = Form(...),
    check_out: str = Form(...),
    guests: int = Form(...)
):
    # Find the hotel
    hotel = next((h for h in hotels if h["id"] == hotel_id), None)

    if not hotel:
        return {"success": False, "message": "Hotel not found"}

    # Create booking
    booking = {
        "id": len(bookings) + 1,
        "hotel": hotel,
        "guest_name": guest_name,
        "email": email,
        "phone": phone,
        "check_in": check_in,
        "check_out": check_out,
        "guests": guests,
        "booking_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    bookings.append(booking)

    # Save to Excel file
    save_booking_to_excel(booking)

    return {
        "success": True,
        "message": f"Booking confirmed for {hotel['name']}!",
        "booking": booking
    }

@app.get("/api/hotels")
async def get_hotels(
    location: Optional[str] = None,
    min_price: Optional[int] = None,
    max_price: Optional[int] = None,
    min_rating: Optional[float] = None,
    amenities: Optional[str] = None,
    hotel_type: Optional[str] = None,
    sort_by: Optional[str] = None
):
    filtered_hotels = hotels.copy()

    # Filter by location
    if location:
        result = []
        for h in filtered_hotels:
            loc = h.get("location")
            if isinstance(loc, str) and location.lower() in loc.lower():
                result.append(h)
        filtered_hotels = result

    # Filter by price range
    if min_price is not None:
        result = []
        for h in filtered_hotels:
            price = h.get("price")
            if isinstance(price, int) and price >= min_price:
                result.append(h)
        filtered_hotels = result

    if max_price is not None:
        result = []
        for h in filtered_hotels:
            price = h.get("price")
            if isinstance(price, int) and price <= max_price:
                result.append(h)
        filtered_hotels = result

    # Filter by rating
    if min_rating is not None:
        result = []
        for h in filtered_hotels:
            rating = h.get("rating")
            if isinstance(rating, (int, float)) and rating >= min_rating:
                result.append(h)
        filtered_hotels = result

    # Filter by amenities
    if amenities:
        amenities_list = [a.strip() for a in amenities.split(",")]
        result = []
        for h in filtered_hotels:
            hotel_amenities = h.get("amenities", [])
            if isinstance(hotel_amenities, list) and all(amenity in hotel_amenities for amenity in amenities_list):
                result.append(h)
        filtered_hotels = result

    # Filter by hotel type
    if hotel_type and hotel_type != "All":
        result = []
        for h in filtered_hotels:
            h_type = h.get("type")
            if h_type == hotel_type:
                result.append(h)
        filtered_hotels = result

    # Sorting
    if sort_by == "price_low":
        filtered_hotels = sorted(filtered_hotels, key=lambda x: x.get("price", 0))
    elif sort_by == "price_high":
        filtered_hotels = sorted(filtered_hotels, key=lambda x: x.get("price", 0), reverse=True)
    elif sort_by == "rating":
        filtered_hotels = sorted(filtered_hotels, key=lambda x: x.get("rating", 0), reverse=True)
    elif sort_by == "popularity":
        filtered_hotels = sorted(filtered_hotels, key=lambda x: x.get("reviews", 0), reverse=True)

    return {"hotels": filtered_hotels, "count": len(filtered_hotels)}

@app.get("/api/cities")
async def get_cities_api():
    return {"cities": get_cities()}

@app.get("/api/amenities")
async def get_amenities_api():
    return {"amenities": get_all_amenities()}

@app.get("/api/bookings")
async def get_bookings():
    """Get all bookings - for demo purposes"""
    return {"bookings": bookings, "count": len(bookings)}

@app.get("/api/download-bookings")
async def download_bookings():
    """Download the Excel file with all bookings"""
    if os.path.exists(EXCEL_FILE):
        return FileResponse(
            path=EXCEL_FILE,
            filename="OYE_Hotel_Bookings.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        return {"success": False, "message": "No bookings file found"}

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 7000))
    uvicorn.run(app, host="0.0.0.0", port=port)
